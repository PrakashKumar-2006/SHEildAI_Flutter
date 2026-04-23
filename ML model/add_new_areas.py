"""
SHEild AI — Add New Bhopal Areas to Dataset
=============================================
Adds 8 new areas:
  Most Important: Ratibad, Patel Nagar, Neelbad, Sakshi Dhaba area, Bhadbhada Road
  Also: DR Ambedkar Nagar, Ayodhya Bypass, Anand Nagar

Then regenerates crime incidents for all new stations and retrains the ML model.
Run: python3 add_new_areas.py
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import random, os, json, warnings
warnings.filterwarnings("ignore")

random.seed(99)

# ══════════════════════════════════════════════════════════════════════════════
# 8 NEW STATIONS — based on real Bhopal geography
# Format: (name, zone, area_type, lat, lon,
#          zone_base, cctv_pct, light_pct, ps_dist_km,
#          isolation, drug, slum, pop_density)
#
# Research basis:
#   Ratibad      — 14km SW Bhopal, semi-rural, Bhadbhada road, very isolated
#   Neelbad      — near Ratibad, limited lighting, no CCTV, high crime
#   Bhadbhada Rd — main road but dark at night, connects isolated areas
#   Sakshi Dhaba — highway area, known for late-night incidents, isolated
#   Patel Nagar  — residential near Raisen Road, moderate risk
#   Anand Nagar  — dense residential, Raisen Road corridor
#   DR Ambedkar  — mixed, eastern Bhopal, moderate infrastructure
#   Ayodhya Byp  — bypass road, isolated, truck route, high night risk
# ══════════════════════════════════════════════════════════════════════════════

NEW_STATIONS = [
    # ── MOST IMPORTANT 5 ──────────────────────────────────────────────────
    ("Ratibad",           "Zone-3 South", "Rural/Peripheral",  23.1452, 77.3580,
     45, 0,  8,  4.2, "Very High", "Yes", "Yes",  "Very Low"),

    ("Neelbad",           "Zone-3 South", "Rural/Peripheral",  23.1580, 77.3720,
     43, 2,  10, 3.8, "Very High", "No",  "Yes",  "Very Low"),

    ("Bhadbhada Road",    "Zone-3 South", "Suburban",          23.1820, 77.3950,
     38, 8,  20, 2.9, "High",      "Yes", "No",   "Low"),

    ("Sakshi Dhaba Area", "Zone-3 South", "Rural/Peripheral",  23.1350, 77.3680,
     46, 0,  5,  5.1, "Very High", "Yes", "No",   "Very Low"),

    ("Patel Nagar",       "Zone-2 East",  "Urban Residential", 23.2480, 77.4580,
     28, 28, 60, 1.4, "Medium",    "No",  "No",   "High"),

    # ── ALSO IMPORTANT 3 ─────────────────────────────────────────────────
    ("Anand Nagar",       "Zone-2 East",  "Urban Residential", 23.2620, 77.4720,
     26, 32, 65, 1.2, "Low",       "No",  "No",   "High"),

    ("DR Ambedkar Nagar", "Zone-2 East",  "Urban Mixed",       23.2180, 77.4650,
     30, 22, 55, 1.6, "Medium",    "No",  "Partial","Medium"),

    ("Ayodhya Bypass",    "Zone-3 South", "Suburban",          23.2040, 77.4760,
     36, 10, 22, 2.8, "High",      "No",  "No",   "Low"),
]

DATASET = "SHEild_AI_Improved_Dataset.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — ADD NEW STATIONS TO SHEET 2
# ══════════════════════════════════════════════════════════════════════════════

print("[1/4] Adding new stations to Sheet 2 (Zone Profile)...")

wb = load_workbook(DATASET)
ws2 = wb["2_Station_Zone_Profile"]

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def dc(ws, row, col, val, alt=False, fx=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin()
    c.fill      = PatternFill("solid", fgColor=fx if fx else ("EEF4FB" if alt else "FFFFFF"))

RC_FILL = {
    "Critical":"FFCCCC","Very High":"FFE0B2",
    "High":"FFF3CD","Medium":"FFF9C4","Low":"D5F5E3"
}

current_last = ws2.max_row  # last data row
start_row = current_last + 1

# Get current max zone_id number
last_zid = ws2.cell(current_last, 1).value  # e.g. "AZ034"
try:
    last_num = int(last_zid.replace("AZ", ""))
except:
    last_num = 34

for i, st in enumerate(NEW_STATIONS):
    sn, zone, atype, lat, lon, zbase, cctv, light, psd, iso, drug, slum, pop = st
    n_avg    = max(8, int(zbase * 1.25))
    overall  = min(100, int(zbase * 1.85))
    rcat     = ("Critical"  if overall >= 82 else
                "Very High" if overall >= 68 else
                "High"      if overall >= 50 else
                "Medium"    if overall >= 35 else "Low")
    route    = ("Avoid"           if overall >= 82 else
                "High Priority"   if overall >= 68 else
                "Medium Priority" if overall >= 50 else "Low Priority")
    trans    = "High" if pop in ["High","Very High"] else "Medium" if pop=="Medium" else "Low"
    zid      = f"AZ{last_num+i+1:03d}"
    cctv_cnt = max(0, int(cctv / 10))

    row_vals = [zid, sn, zone, atype, lat, lon, cctv, light, psd,
                drug, slum, iso, pop, n_avg, zbase, overall, rcat, route, trans, cctv_cnt]
    ri       = start_row + i
    alt      = ri % 2 == 0
    for ci, val in enumerate(row_vals, 1):
        fx = RC_FILL.get(rcat) if ci == 17 else None
        dc(ws2, ri, ci, val, alt=alt, fx=fx)

print(f"  Added {len(NEW_STATIONS)} new stations (AZ035–AZ{last_num+len(NEW_STATIONS):03d})")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — GENERATE CRIME INCIDENTS FOR NEW STATIONS → ADD TO SHEET 1
# ══════════════════════════════════════════════════════════════════════════════

print("[2/4] Generating crime incidents for new stations...")

ws1 = wb["1_Crime_Incidents"]
current_last1 = ws1.max_row
last_iid_val  = ws1.cell(current_last1, 1).value
try:
    iid = int(last_iid_val) + 1
except:
    iid = 9001

CRIMES = [
    ("Rape",                 10,"IPC 376",  "Sexual Violence",  "18-30","Known/Acquaintance",["Residential","Isolated"],        ["Night","Late Evening"]),
    ("Gang Rape",            10,"IPC 376D", "Sexual Violence",  "18-25","Stranger Group",    ["Isolated","Vehicle"],            ["Night"]),
    ("Molestation/Assault",   7,"IPC 354",  "Sexual Harassment","18-30","Stranger",          ["Public","Street","Bus Stop"],   ["Evening","Late Evening"]),
    ("Sexual Harassment",     6,"IPC 354A", "Sexual Harassment","18-35","Stranger",          ["Workplace","Public"],           ["Morning","Afternoon"]),
    ("Stalking",              5,"IPC 354D", "Sexual Harassment","18-30","Known/Stranger",    ["Street","Near Home"],           ["Any"]),
    ("Kidnapping/Abduction",  8,"IPC 363",  "Abduction",        "Under 18","Stranger",       ["School Zone","Open Area"],      ["Afternoon","Evening"]),
    ("Abduction of Women",    8,"IPC 364",  "Abduction",        "18-30","Known/Stranger",    ["Street","Isolated"],            ["Evening","Night"]),
    ("Cruelty by Husband",    5,"IPC 498A", "Domestic Violence","26-40","Husband/In-laws",   ["Domestic"],                    ["Morning","Any"]),
    ("Dowry Death",           9,"IPC 304B", "Domestic Violence","26-35","Husband/In-laws",   ["Domestic"],                    ["Morning","Afternoon"]),
    ("Dowry Harassment",      4,"IPC 498A", "Domestic Violence","26-40","Husband/In-laws",   ["Domestic"],                    ["Any"]),
    ("Domestic Violence",     6,"PWDVA",    "Domestic Violence","26-45","Husband/Partner",   ["Domestic"],                    ["Any"]),
    ("Attempt to Murder",     9,"IPC 307",  "Violent Crime",    "Any",  "Known/Stranger",    ["Any"],                         ["Night","Late Evening"]),
    ("Acid Attack",          10,"IPC 326A", "Violent Crime",    "18-30","Rejected Suitor",   ["Street","Near Home"],          ["Any"]),
    ("Cyber Crime/Stalking",  3,"IT Act 67","Cyber Crime",      "18-35","Online/Unknown",    ["Online"],                      ["Any"]),
    ("Human Trafficking",    10,"IPC 370",  "Trafficking",      "Under 18","Organised Group",["Isolated","Station"],          ["Any"]),
    ("POCSO Offense",        10,"POCSO 4",  "Child Crime",      "Under 18","Known/Family",   ["Home","School"],               ["Any"]),
    ("Insult to Modesty",     4,"IPC 509",  "Sexual Harassment","Any",  "Stranger",          ["Public","Workplace"],          ["Any"]),
]
CRIME_W = [0.05,0.01,0.17,0.07,0.05,0.11,0.05,0.21,0.02,0.07,0.04,0.02,0.005,0.04,0.01,0.02,0.02]
CRIME_W_NORM = [w/sum(CRIME_W) for w in CRIME_W]

TIME_SLOTS = [
    ("Early Morning (00-06)", 0, 6,  16),
    ("Morning (06-10)",       6, 10,  0),
    ("Late Morning (10-12)",  10,12, -6),
    ("Afternoon (12-16)",     12,16, -2),
    ("Late Afternoon (16-18)",16,18,  2),
    ("Evening (18-21)",       18,21, 10),
    ("Late Evening (21-24)",  21,24, 14),
]
SLOT_W = [0.08,0.12,0.07,0.11,0.14,0.23,0.18]
SLOT_W_NORM = [w/sum(SLOT_W) for w in SLOT_W]

MONTHS   = ["January","February","March","April","May","June",
            "July","August","September","October","November","December"]
MONTH_W  = [1.0,1.0,1.0,1.1,1.2,1.0,1.0,1.1,1.1,1.2,1.3,1.2]
MONTH_WN = [w/sum(MONTH_W) for w in MONTH_W]
SEASONS  = {"January":"Winter","February":"Winter","March":"Spring",
            "April":"Summer","May":"Summer","June":"Monsoon",
            "July":"Monsoon","August":"Monsoon","September":"Post-Monsoon",
            "October":"Post-Monsoon","November":"Winter","December":"Winter"}
WEATHER  = {"Winter":["Cold/Foggy","Clear","Partly Cloudy"],
            "Spring":["Clear","Partly Cloudy","Warm"],
            "Summer":["Hot","Very Hot","Partly Cloudy"],
            "Monsoon":["Rainy","Heavy Rain","Overcast"],
            "Post-Monsoon":["Clear","Partly Cloudy","Mild"]}
DAYS     = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
VICTIM_AGE = ["Under 18","18-25","26-30","31-40","41-50","Above 50"]
VICTIM_W   = [0.15,0.30,0.20,0.20,0.10,0.05]
RELATIONS  = ["Husband/In-laws","Known/Acquaintance","Stranger",
              "Online/Unknown","Family Member","Neighbor","Employer/Colleague"]
LOCATIONS  = ["Domestic/Home","Public Street","Isolated Road","Market/Bazaar",
              "Public Transport","Near School/College","Park/Open Ground",
              "Workplace","Railway/Bus Station","Slum/Colony","Agricultural Field"]
CHARGESHEET = ["Chargesheeted","Pending Investigation","Closed-Insufficient Evidence",
               "Conviction","Acquittal"]
CS_W = [0.55,0.30,0.08,0.05,0.02]

def pick(items, weights):
    r, cum = random.random(), 0.0
    for item, w in zip(items, weights):
        cum += w
        if r <= cum: return item
    return items[-1]

def compute_risk(zbase, severity, hr, cctv_yn, lighting, psd, isolation, drug, is_wknd):
    score   = float(zbase)
    sl      = next((s for s in TIME_SLOTS if s[1] <= hr < s[2]), TIME_SLOTS[0])
    score  += sl[3]
    score  += severity * 1.5
    score  += {"Good":0,"Moderate":4,"Poor":9,"None":16,"N/A":2}.get(lighting, 4)
    cctv_n  = {"Yes":1.0,"Partial":0.5,"No":0.0}.get(cctv_yn, 0)
    score  -= cctv_n * 12
    score  += {"Very High":18,"High":12,"Medium":6,"Low":0,"No":0}.get(isolation, 4)
    score  += 8 if drug == "Yes" else 0
    score  += 0 if psd < 1 else 5 if psd < 2 else 10 if psd < 3 else 14
    score  += 4 if is_wknd else 0
    return max(0.0, min(100.0, round(score, 1)))

new_rows = []

for year in [2020, 2021, 2022, 2023]:
    for st in NEW_STATIONS:
        sn, zone, atype, lat, lon, zbase, cctv_pct, light_pct, psd, iso, drug, slum, pop = st
        n = max(8, int(zbase * 1.25))

        for _ in range(n):
            ct  = pick(CRIMES, CRIME_W_NORM)
            sl  = pick(TIME_SLOTS, SLOT_W_NORM)
            hr  = random.randint(sl[1], max(sl[1], sl[2]-1))
            is_night  = int(hr >= 21 or hr < 6)
            is_eve    = int(18 <= hr < 21)
            is_day    = int(10 <= hr < 16)
            mo        = pick(MONTHS, MONTH_WN)
            mo_num    = MONTHS.index(mo) + 1
            dy        = random.choice(DAYS)
            is_wknd   = int(dy in ["Saturday","Sunday"])
            season    = SEASONS[mo]
            weather   = random.choice(WEATHER[season])

            if is_night:
                lighting = pick(["Good","Moderate","Poor","None"], [0.05,0.15,0.45,0.35])
            else:
                lighting = pick(["Good","Moderate","Poor","None"], [0.45,0.35,0.15,0.05])

            cp      = cctv_pct / 100.0
            cctv_yn = pick(["Yes","Partial","No"],
                           [max(0,cp*0.8), min(0.2, 0.3-cp*0.2), max(0.05, 1-cp*0.8-min(0.2,0.3-cp*0.2))])

            if "26-" in ct[4] or "26-45" in ct[4]:
                v_age = pick(["26-30","31-40","41-50"], [0.40,0.45,0.15])
            else:
                v_age = pick(VICTIM_AGE, VICTIM_W)

            if "Husband" in ct[5]:
                relation = "Husband/In-laws"
            elif "Known" in ct[5]:
                relation = pick(["Known/Acquaintance","Neighbor","Family Member"], [0.6,0.25,0.15])
            elif "Online" in str(ct[6]):
                relation = "Online/Unknown"
            else:
                relation = pick(RELATIONS, [0.20,0.25,0.30,0.10,0.05,0.05,0.05])

            locs = ct[6]
            if "Domestic" in locs[0]:
                loc = "Domestic/Home"
            elif "Online" in locs[0]:
                loc = "Online/Cyber"
            elif "Isolated" in locs[0]:
                loc = pick(["Isolated Road","Agricultural Field","Forest/Outskirts"], [0.5,0.3,0.2])
            elif "School" in locs[0]:
                loc = pick(["Near School/College","Public Street","Park/Open Ground"], [0.5,0.3,0.2])
            else:
                loc = pick(LOCATIONS[:8], [0.10,0.20,0.10,0.18,0.12,0.10,0.10,0.10])

            rs = compute_risk(zbase, ct[1], hr, cctv_yn, lighting, psd, iso, drug, is_wknd)
            rl = 1 if rs <= 35 else 2 if rs <= 62 else 3
            rc = {1:"Medium",2:"High",3:"Critical"}[rl]
            cs = pick(CHARGESHEET, CS_W)

            jlat = round(lat + random.uniform(-0.010, 0.010), 6)
            jlon = round(lon + random.uniform(-0.010, 0.010), 6)

            new_rows.append([
                iid, year, mo, mo_num, dy, is_wknd, season, weather,
                sn, zone, atype, jlat, jlon,
                ct[0], ct[2], ct[3], ct[1],
                sl[0], hr, is_night, is_eve, is_day,
                v_age, relation, loc,
                cctv_yn, lighting, psd, iso, drug,
                cs, rs, rl, rc
            ])
            iid += 1

print(f"  Generated {len(new_rows):,} new crime incidents")

# Append to Sheet 1
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CRIME_FILL = {
    "Sexual Violence":"FFCCCC","Sexual Harassment":"FFE0CC","Abduction":"FFE5B4",
    "Domestic Violence":"FFF3CD","Violent Crime":"F4CCCC","Cyber Crime":"D9EAD3",
    "Trafficking":"F9CBFF","Child Crime":"FFD9D9",
}

for ri_off, row in enumerate(new_rows):
    ri  = current_last1 + 1 + ri_off
    alt = ri % 2 == 0
    cat = row[15]
    rl  = row[32]
    for ci, val in enumerate(row, 1):
        c             = ws1.cell(row=ri, column=ci, value=val)
        c.font        = Font(name="Arial", size=9)
        c.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border      = thin()
        fx = (CRIME_FILL.get(cat) if ci == 16
              else {1:"D5F5E3",2:"FEF9E7",3:"FADBD8"}.get(rl) if ci == 33
              else None)
        c.fill = PatternFill("solid", fgColor=fx if fx else ("EEF4FB" if alt else "FFFFFF"))

wb.save(DATASET)
print(f"  Dataset saved with {4856 + len(new_rows):,} total incidents")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — RETRAIN ML MODEL ON FULL UPDATED DATASET
# ══════════════════════════════════════════════════════════════════════════════

print("\n[3/4] Retraining ML model on full updated dataset...")

from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, VotingClassifier
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics import accuracy_score, f1_score, classification_report
import joblib

os.makedirs("models", exist_ok=True)

df_inc  = pd.read_excel(DATASET, sheet_name="1_Crime_Incidents",          header=1)
df_zone = pd.read_excel(DATASET, sheet_name="2_Station_Zone_Profile",     header=1)
df_time = pd.read_excel(DATASET, sheet_name="3_Time_Environment_Factors", header=1)

print(f"  Loaded {len(df_inc):,} incidents, {len(df_zone)} stations")

df = df_inc.copy()
df["hour_24"]        = df["Hour_24"].astype(float)
df["crime_severity"] = df["Crime_Severity_Score"].astype(float)
df["police_dist_km"] = df["Nearest_PS_Dist_km"].astype(float)
df["is_weekend"]     = df["Is_Weekend"].astype(int)
df["cctv_numeric"]   = df["CCTV_Available"].str.strip().map({"Yes":1.0,"Partial":0.5,"No":0.0}).fillna(0.0)
df["lighting_numeric"] = df["Street_Lighting"].str.strip().map({"Good":0,"Moderate":1,"Poor":2,"None":3,"N/A":1}).fillna(1)
df["is_night"]   = ((df["hour_24"]>=21)|(df["hour_24"]<6)).astype(int)
df["is_evening"] = ((df["hour_24"]>=18)&(df["hour_24"]<21)).astype(int)
df["is_daytime"] = ((df["hour_24"]>=10)&(df["hour_24"]<16)).astype(int)
df["hour_sin"]   = np.sin(2*np.pi*df["hour_24"]/24)
df["hour_cos"]   = np.cos(2*np.pi*df["hour_24"]/24)
df["month_sin"]  = np.sin(2*np.pi*df["Month_Num"]/12)
df["month_cos"]  = np.cos(2*np.pi*df["Month_Num"]/12)

def get_time_feats(h):
    for _, row in df_time.iterrows():
        try:
            s, e = [int(x) for x in str(row["Hour_Range"]).split("-")]
            if s <= h < e:
                return float(row["Time_Additive_Score"]), float(row["Incident_Share_Pct"])
        except: pass
    return 0.0, 10.0

df["time_additive"], df["incident_share"] = zip(*df["hour_24"].apply(get_time_feats))

df_zone2 = df_zone[["Police_Station","Zone_Base_Score","Overall_Risk_Score_Display",
                     "CCTV_Coverage_Pct","Nearest_PS_Dist_km",
                     "Drug_Alcohol_Nearby","Slum_Area","Isolated_Roads","Population_Density"]].copy()
df_zone2["drug_nearby"] = (df_zone2["Drug_Alcohol_Nearby"].str.lower()=="yes").astype(int)
df_zone2["slum_area"]   = df_zone2["Slum_Area"].str.lower().map({"yes":1,"partial":0.5,"no":0}).fillna(0)
df_zone2["isolation"]   = df_zone2["Isolated_Roads"].str.lower().map({"very high":4,"high":3,"medium":2,"low":1,"no":0}).fillna(1)
df_zone2["pop_density"] = df_zone2["Population_Density"].str.lower().str.extract(r"(very high|high|medium|low|very low)")[0].map({"very high":5,"high":4,"medium":3,"low":2,"very low":1}).fillna(2)
df_zone2["zone_base"]    = pd.to_numeric(df_zone2["Zone_Base_Score"],            errors="coerce").fillna(25)
df_zone2["zone_risk"]    = pd.to_numeric(df_zone2["Overall_Risk_Score_Display"], errors="coerce").fillna(50)
df_zone2["zone_cctv"]    = pd.to_numeric(df_zone2["CCTV_Coverage_Pct"],          errors="coerce").fillna(20)
df_zone2["zone_ps_dist"] = pd.to_numeric(df_zone2["Nearest_PS_Dist_km"],         errors="coerce").fillna(2)
zs = df_zone2[["Police_Station","zone_base","zone_risk","zone_cctv","zone_ps_dist",
               "drug_nearby","slum_area","isolation","pop_density"]].drop_duplicates("Police_Station")

df = df.merge(zs, on="Police_Station", how="left")
zcols = ["zone_base","zone_risk","zone_cctv","zone_ps_dist","drug_nearby","slum_area","isolation","pop_density"]
df[zcols] = df[zcols].fillna(df[zcols].median())

le_crime   = LabelEncoder(); df["crime_enc"]   = le_crime.fit_transform(df["Crime_Type"].astype(str))
le_zone    = LabelEncoder(); df["zone_enc"]    = le_zone.fit_transform(df["Zone"].astype(str))
le_area    = LabelEncoder(); df["area_enc"]    = le_area.fit_transform(df["Area_Type"].astype(str))
le_victim  = LabelEncoder(); df["victim_enc"]  = le_victim.fit_transform(df["Victim_Age_Group"].astype(str))
le_rel     = LabelEncoder(); df["rel_enc"]     = le_rel.fit_transform(df["Relation_to_Accused"].astype(str))
le_loc     = LabelEncoder(); df["loc_enc"]     = le_loc.fit_transform(df["Location_Context"].astype(str))
le_weather = LabelEncoder(); df["weather_enc"] = le_weather.fit_transform(df["Weather"].astype(str))

df["night_isolation"]    = df["is_night"]       * df["isolation"]
df["night_no_cctv"]      = df["is_night"]       * (1 - df["cctv_numeric"])
df["evening_isolation"]  = df["is_evening"]     * df["isolation"]
df["severity_x_night"]   = df["crime_severity"] * df["is_night"]
df["zone_risk_x_time"]   = df["zone_base"]      * df["time_additive"]
df["dist_x_isolation"]   = df["zone_ps_dist"]   * df["isolation"]
df["weekend_x_night"]    = df["is_weekend"]     * df["is_night"]
df["severity_x_evening"] = df["crime_severity"] * df["is_evening"]

FEATURES = [
    "hour_24","crime_severity","police_dist_km","cctv_numeric","lighting_numeric",
    "is_night","is_evening","is_daytime","is_weekend",
    "hour_sin","hour_cos","month_sin","month_cos",
    "time_additive","incident_share",
    "zone_base","zone_risk","zone_cctv","zone_ps_dist",
    "drug_nearby","slum_area","isolation","pop_density",
    "crime_enc","zone_enc","area_enc","victim_enc","rel_enc","loc_enc","weather_enc",
    "night_isolation","night_no_cctv","evening_isolation",
    "severity_x_night","zone_risk_x_time","dist_x_isolation",
    "weekend_x_night","severity_x_evening",
]

X = df[FEATURES].fillna(0)
y = df["Risk_Label"]
print(f"  Feature matrix: {X.shape}")
print(f"  Label dist: {dict(y.value_counts().sort_index())}")

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20, random_state=42, stratify=y)
scaler    = StandardScaler()
X_train_s = scaler.fit_transform(X_train)
X_test_s  = scaler.transform(X_test)

rf = RandomForestClassifier(n_estimators=400, max_depth=14, min_samples_split=2,
     min_samples_leaf=1, max_features="sqrt", class_weight="balanced", random_state=42, n_jobs=-1)
rf.fit(X_train_s, y_train)

gb = GradientBoostingClassifier(n_estimators=250, learning_rate=0.08,
     max_depth=5, subsample=0.85, random_state=42)
gb.fit(X_train_s, y_train)

ensemble = VotingClassifier(estimators=[("rf",rf),("gb",gb)], voting="soft", weights=[2,1])
ensemble.fit(X_train_s, y_train)
ens_acc = accuracy_score(y_test, ensemble.predict(X_test_s))
ens_f1  = f1_score(y_test, ensemble.predict(X_test_s), average="weighted")

cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
cv_scores = cross_val_score(ensemble, scaler.transform(X), y, cv=cv, scoring="f1_weighted")

print(f"  Accuracy : {ens_acc*100:.2f}%")
print(f"  F1 Score : {ens_f1:.3f}")
print(f"  CV F1    : {cv_scores.mean():.3f} ± {cv_scores.std():.4f}")
print()
print("  Classification Report:")
print(classification_report(y_test, ensemble.predict(X_test_s),
      target_names=["Medium(1)","High(2)","Critical(3)"]))

feat_imp = dict(zip(FEATURES, rf.feature_importances_))

bundle = {
    "model": ensemble, "scaler": scaler, "features": FEATURES,
    "encoders": {
        "crime_type":le_crime,"zone":le_zone,"area_type":le_area,
        "victim_age":le_victim,"relation":le_rel,"location":le_loc,"weather":le_weather,
    },
    "zone_lookup": zs.set_index("Police_Station").to_dict(orient="index"),
    "metrics": {
        "accuracy":         round(ens_acc, 4),
        "f1_weighted":      round(ens_f1, 4),
        "cv_f1_mean":       round(float(cv_scores.mean()), 4),
        "cv_f1_std":        round(float(cv_scores.std()), 4),
        "training_records": len(X_train),
        "test_records":     len(X_test),
        "total_features":   len(FEATURES),
        "total_stations":   len(df_zone),
        "new_stations_added": [s[0] for s in NEW_STATIONS],
    },
    "label_map":        {1:"MEDIUM",2:"HIGH",3:"CRITICAL"},
    "label_thresholds": {"MEDIUM":"0-35","HIGH":"36-62","CRITICAL":"63-100"},
    "feature_importance": feat_imp,
}

joblib.dump(bundle, "models/sheild_risk_model.pkl")
with open("models/model_metrics.json","w") as f:
    json.dump(bundle["metrics"], f, indent=2)

print("\n[4/4] Done!")
print("="*65)
print(f"  Dataset  : {4856+len(new_rows):,} incidents  |  {len(df_zone)} stations")
print(f"  Accuracy : {ens_acc*100:.2f}%")
print(f"  CV F1    : {cv_scores.mean():.3f} ± {cv_scores.std():.4f}")
print()
print("  New areas now covered by ML model:")
for st in NEW_STATIONS:
    overall = min(100, int(st[5]*1.85))
    rcat = "Critical" if overall>=82 else "Very High" if overall>=68 else "High" if overall>=50 else "Medium"
    print(f"    {st[0]:<22} → Risk: {overall}/100  ({rcat})")
print()
print("  Next: uvicorn main:app --reload --host 0.0.0.0 --port 8000")
print("="*65)
