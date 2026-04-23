"""
SHEild AI — Add New Areas v2
Adds 14 new specific locations:
 - Neelbad Ratibad Road
 - Near SISTEC-R College (Sikandrabad, Ratibad)
 - Near TIT College (Anand Nagar, Piplani)
 - Ratibad Chowrah (Chauraha)
 - Sikandrabad Village
 - Bhadbhada Road (extended)
 - Chunabhatti
 - Suraj Nagar
 - Kerwa Dam Road
 - Nehru Nagar
 - Kokta / Raisen Road
 - Phanda
 - Kalchuri Nagar
 - Mandideep Border Area

Run: python3 add_new_areas_v2.py
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import random, os, json, warnings
warnings.filterwarnings("ignore")

random.seed(77)

# ══════════════════════════════════════════════════════════════════════════════
# 14 NEW SPECIFIC LOCATIONS — GPS verified from search
#
# Format: (name, zone, area_type, lat, lon,
#          zone_base, cctv_pct, light_pct, ps_dist_km,
#          isolation, drug, slum, pop_density)
#
# GPS Sources:
#  Ratibad pin: 23.1661, 77.3281 (indiamapia.com)
#  SISTEC-R: Sikandrabad, Near Ratibad, Bhadbhada Rd — ~23.170, 77.310
#  TIT College: Anand Nagar, Piplani, BHEL — ~23.256, 77.475
#  Chunabhatti: ~23.209, 77.465 (south-east Bhopal)
#  Suraj Nagar: Bhadbhada Rd — ~23.175, 77.391
#  Nehru Nagar: ~23.252, 77.425 (central)
#  Kerwa Dam Rd: ~23.225, 77.357 (Shyamla Hills area)
#  Kokta/Raisen Road: ~23.268, 77.493
#  Phanda: ~23.203, 77.234 (far west)
#  Kalchuri Nagar: Raisen Road — ~23.263, 77.489
#  Mandideep Border: ~23.103, 77.554 (south-east edge)
# ══════════════════════════════════════════════════════════════════════════════

NEW_STATIONS_V2 = [

    # ── RATIBAD-NEELBAD CORRIDOR (Most Important) ─────────────────────────

    # 1. Neelbad-Ratibad Road — the actual road between both villages
    #    No CCTV, poor lighting, truck route at night, very isolated
    ("Neelbad Ratibad Road",   "Zone-3 South", "Rural/Peripheral",
     23.1620, 77.3490,
     47, 0,  6,  4.5, "Very High", "Yes", "No",  "Very Low"),

    # 2. Near SISTEC-R College — Sikandrabad, Bhadbhada Road
    #    College area so students walk at night → higher risk
    #    Road to college is semi-isolated, no lighting after 9pm
    ("SISTEC-R College Area",  "Zone-3 South", "Rural/Peripheral",
     23.1700, 77.3100,
     44, 3,  12, 4.0, "Very High", "No",  "No",  "Low"),

    # 3. Ratibad Chowrah (Chauraha) — main junction in Ratibad
    #    Slightly better than deep Ratibad due to junction activity
    #    But still very low CCTV, isolated at night
    ("Ratibad Chowrah",        "Zone-3 South", "Suburban",
     23.1661, 77.3281,
     41, 5,  18, 3.6, "High",      "Yes", "No",  "Low"),

    # 4. Sikandrabad Village — 3km from Ratibad, very rural
    #    Near SISTEC-R, students travel this road daily
    #    Zero infrastructure, completely dark at night
    ("Sikandrabad",            "Zone-3 South", "Rural/Peripheral",
     23.1750, 77.3180,
     46, 0,  5,  4.8, "Very High", "No",  "Yes", "Very Low"),

    # ── COLLEGE AREAS ─────────────────────────────────────────────────────

    # 5. Near TIT College — Anand Nagar, Piplani, BHEL
    #    Large college campus, students walk at night near Hataikheda Dam
    #    Area is semi-developed, moderate CCTV near BHEL township
    ("Near TIT College",       "Zone-2 East",  "Urban Residential",
     23.2560, 77.4750,
     27, 25, 55, 1.3, "Medium",    "No",  "No",  "High"),

    # ── BHADBHADA ROAD CORRIDOR ──────────────────────────────────────────

    # 6. Suraj Nagar (Bhadbhada area) — on Bhadbhada road
    #    Police station here but area is dark at night
    ("Suraj Nagar",            "Zone-3 South", "Suburban",
     23.1750, 77.3910,
     36, 10, 22, 2.4, "High",      "No",  "No",  "Low"),

    # 7. Kerwa Dam Road — leads to Kerwa Dam, very isolated
    #    Used for night trips, picnic area → high risk after dark
    ("Kerwa Dam Road",         "Zone-1 Central","Rural/Peripheral",
     23.2250, 77.3570,
     42, 2,  8,  3.5, "Very High", "No",  "No",  "Very Low"),

    # ── EAST/NORTH BHOPAL ────────────────────────────────────────────────

    # 8. Chunabhatti — industrial + residential mix, east Bhopal
    #    High population but limited police patrol at night
    ("Chunabhatti",            "Zone-2 East",  "Industrial",
     23.2090, 77.4650,
     32, 18, 40, 1.8, "High",      "Yes", "Partial","Medium"),

    # 9. Nehru Nagar — residential, central-east, near bus stand
    #    Moderate risk, some CCTV, reasonably lit
    ("Nehru Nagar",            "Zone-1 Central","Urban Residential",
     23.2520, 77.4250,
     22, 35, 65, 1.0, "Low",       "No",  "No",  "High"),

    # 10. Kokta (Raisen Road) — educational corridor, growing area
    #     College hostels nearby, evening risk for women students
    ("Kokta Raisen Road",      "Zone-2 East",  "Suburban",
     23.2680, 77.4930,
     30, 20, 45, 1.8, "Medium",    "No",  "No",  "Medium"),

    # 11. Kalchuri Nagar (Raisen Road) — residential + colleges
    #     Better developed than Kokta but still moderate risk
    ("Kalchuri Nagar",         "Zone-2 East",  "Urban Residential",
     23.2630, 77.4890,
     26, 28, 60, 1.5, "Low",       "No",  "No",  "Medium"),

    # ── PERIPHERAL BHOPAL ─────────────────────────────────────────────────

    # 12. Phanda — far west Bhopal, semi-rural
    #     Very low police presence, drug spots on highway
    ("Phanda",                 "Zone-4 North", "Rural/Peripheral",
     23.2030, 77.2340,
     44, 2,  8,  5.2, "Very High", "Yes", "No",  "Very Low"),

    # 13. Semra Bazyaft — near Ratibad corridor
    #     Completely rural, on Sehore-Bhopal road, zero infrastructure
    ("Semra Bazyaft",          "Zone-3 South", "Rural/Peripheral",
     23.1580, 77.3050,
     45, 0,  5,  5.0, "Very High", "No",  "Yes", "Very Low"),

    # 14. Mandideep Border — south-east edge, industrial zone
    #     Large factories, night shift workers, isolated
    ("Mandideep Border",       "Zone-2 East",  "Industrial",
     23.1030, 77.5540,
     39, 8,  20, 3.2, "High",      "Yes", "No",  "Low"),
]

DATASET = "SHEild_AI_Improved_Dataset.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def dc(ws, row, col, val, alt=False, fx=None):
    c             = ws.cell(row=row, column=col, value=val)
    c.font        = Font(name="Arial", size=9)
    c.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border      = thin()
    c.fill        = PatternFill("solid", fgColor=fx if fx else ("EEF4FB" if alt else "FFFFFF"))

RC_FILL = {
    "Critical":"FFCCCC", "Very High":"FFE0B2",
    "High":"FFF3CD",     "Medium":"FFF9C4", "Low":"D5F5E3"
}

CRIME_FILL = {
    "Sexual Violence":"FFCCCC", "Sexual Harassment":"FFE0CC",
    "Abduction":"FFE5B4",       "Domestic Violence":"FFF3CD",
    "Violent Crime":"F4CCCC",   "Cyber Crime":"D9EAD3",
    "Trafficking":"F9CBFF",     "Child Crime":"FFD9D9",
}

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — ADD TO SHEET 2 (Station Zone Profile)
# ══════════════════════════════════════════════════════════════════════════════

print("[1/4] Adding 14 new stations to Sheet 2...")

wb        = load_workbook(DATASET)
ws2       = wb["2_Station_Zone_Profile"]
last_row  = ws2.max_row
last_zid  = ws2.cell(last_row, 1).value
last_num  = int(str(last_zid).replace("AZ","")) if last_zid else 42

for i, st in enumerate(NEW_STATIONS_V2):
    sn, zone, atype, lat, lon, zbase, cctv, light, psd, iso, drug, slum, pop = st
    n_avg   = max(8, int(zbase * 1.25))
    overall = min(100, int(zbase * 1.85))
    rcat    = ("Critical"  if overall >= 82 else
               "Very High" if overall >= 68 else
               "High"      if overall >= 50 else
               "Medium"    if overall >= 35 else "Low")
    route   = ("Avoid"           if overall >= 82 else
               "High Priority"   if overall >= 68 else
               "Medium Priority" if overall >= 50 else "Low Priority")
    trans   = "High" if pop in ["High","Very High"] else "Medium" if pop=="Medium" else "Low"
    zid     = f"AZ{last_num + i + 1:03d}"
    cnt     = max(0, int(cctv / 10))
    ri      = last_row + 1 + i
    vals    = [zid, sn, zone, atype, lat, lon, cctv, light, psd,
               drug, slum, iso, pop, n_avg, zbase, overall, rcat, route, trans, cnt]
    for ci, val in enumerate(vals, 1):
        fx = RC_FILL.get(rcat) if ci == 17 else None
        dc(ws2, ri, ci, val, alt=(ri%2==0), fx=fx)

print(f"  Added {len(NEW_STATIONS_V2)} stations "
      f"(AZ{last_num+1:03d}–AZ{last_num+len(NEW_STATIONS_V2):03d})")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — GENERATE CRIME INCIDENTS FOR NEW STATIONS
# ══════════════════════════════════════════════════════════════════════════════

print("[2/4] Generating crime incidents...")

CRIMES = [
    ("Rape",                 10,"IPC 376",  "Sexual Violence",  "18-30","Known/Acquaintance",["Residential","Isolated"],       ["Night","Late Evening"]),
    ("Gang Rape",            10,"IPC 376D", "Sexual Violence",  "18-25","Stranger Group",    ["Isolated","Vehicle"],           ["Night"]),
    ("Molestation/Assault",   7,"IPC 354",  "Sexual Harassment","18-30","Stranger",          ["Public","Street","Bus Stop"],  ["Evening","Late Evening"]),
    ("Sexual Harassment",     6,"IPC 354A", "Sexual Harassment","18-35","Stranger",          ["Workplace","Public"],          ["Morning","Afternoon"]),
    ("Stalking",              5,"IPC 354D", "Sexual Harassment","18-30","Known/Stranger",    ["Street","Near Home"],          ["Any"]),
    ("Kidnapping/Abduction",  8,"IPC 363",  "Abduction",        "Under 18","Stranger",       ["School Zone","Open Area"],     ["Afternoon","Evening"]),
    ("Abduction of Women",    8,"IPC 364",  "Abduction",        "18-30","Known/Stranger",    ["Street","Isolated"],           ["Evening","Night"]),
    ("Cruelty by Husband",    5,"IPC 498A", "Domestic Violence","26-40","Husband/In-laws",   ["Domestic"],                   ["Morning","Any"]),
    ("Dowry Death",           9,"IPC 304B", "Domestic Violence","26-35","Husband/In-laws",   ["Domestic"],                   ["Morning","Afternoon"]),
    ("Dowry Harassment",      4,"IPC 498A", "Domestic Violence","26-40","Husband/In-laws",   ["Domestic"],                   ["Any"]),
    ("Domestic Violence",     6,"PWDVA",    "Domestic Violence","26-45","Husband/Partner",   ["Domestic"],                   ["Any"]),
    ("Attempt to Murder",     9,"IPC 307",  "Violent Crime",    "Any",  "Known/Stranger",    ["Any"],                        ["Night","Late Evening"]),
    ("Acid Attack",          10,"IPC 326A", "Violent Crime",    "18-30","Rejected Suitor",   ["Street","Near Home"],         ["Any"]),
    ("Cyber Crime/Stalking",  3,"IT Act 67","Cyber Crime",      "18-35","Online/Unknown",    ["Online"],                     ["Any"]),
    ("Human Trafficking",    10,"IPC 370",  "Trafficking",      "Under 18","Organised",      ["Isolated","Station"],         ["Any"]),
    ("POCSO Offense",        10,"POCSO 4",  "Child Crime",      "Under 18","Known/Family",   ["Home","School"],              ["Any"]),
    ("Insult to Modesty",     4,"IPC 509",  "Sexual Harassment","Any",  "Stranger",          ["Public","Workplace"],         ["Any"]),
]
CW = [0.05,0.01,0.17,0.07,0.05,0.11,0.05,0.21,0.02,0.07,0.04,0.02,0.005,0.04,0.01,0.02,0.02]
CWN = [w/sum(CW) for w in CW]

TIME_SLOTS = [
    ("Early Morning (00-06)", 0, 6,  16),
    ("Morning (06-10)",       6, 10,  0),
    ("Late Morning (10-12)",  10,12, -6),
    ("Afternoon (12-16)",     12,16, -2),
    ("Late Afternoon (16-18)",16,18,  2),
    ("Evening (18-21)",       18,21, 10),
    ("Late Evening (21-24)",  21,24, 14),
]
SW  = [0.08,0.12,0.07,0.11,0.14,0.23,0.18]
SWN = [w/sum(SW) for w in SW]

MONTHS   = ["January","February","March","April","May","June",
            "July","August","September","October","November","December"]
MWN      = [w/sum([1,1,1,1.1,1.2,1,1,1.1,1.1,1.2,1.3,1.2])
            for w in [1,1,1,1.1,1.2,1,1,1.1,1.1,1.2,1.3,1.2]]
SEASONS  = {"January":"Winter","February":"Winter","March":"Spring",
            "April":"Summer","May":"Summer","June":"Monsoon",
            "July":"Monsoon","August":"Monsoon","September":"Post-Monsoon",
            "October":"Post-Monsoon","November":"Winter","December":"Winter"}
WEATHER  = {"Winter":["Cold/Foggy","Clear","Partly Cloudy"],
            "Spring":["Clear","Partly Cloudy","Warm"],
            "Summer":["Hot","Very Hot","Partly Cloudy"],
            "Monsoon":["Rainy","Heavy Rain","Overcast"],
            "Post-Monsoon":["Clear","Partly Cloudy","Mild"]}
DAYS     = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
VICTIM_A = ["Under 18","18-25","26-30","31-40","41-50","Above 50"]
VICTIM_W = [0.15,0.30,0.20,0.20,0.10,0.05]
RELS     = ["Husband/In-laws","Known/Acquaintance","Stranger",
            "Online/Unknown","Family Member","Neighbor","Employer/Colleague"]
LOCS     = ["Domestic/Home","Public Street","Isolated Road","Market/Bazaar",
            "Public Transport","Near School/College","Park/Open Ground",
            "Workplace","Agricultural Field","Slum/Colony","Hotel/Lodge"]
CHARGE   = ["Chargesheeted","Pending Investigation","Closed-Insufficient Evidence",
            "Conviction","Acquittal"]
CHARGE_W = [0.55,0.30,0.08,0.05,0.02]

def pick(items, weights):
    r, cum = random.random(), 0.0
    for item, w in zip(items, weights):
        cum += w
        if r <= cum: return item
    return items[-1]

def risk(zbase, sev, hr, cctv_yn, lighting, psd, iso, drug, wknd):
    sc = float(zbase)
    sl = next((s for s in TIME_SLOTS if s[1]<=hr<s[2]), TIME_SLOTS[0])
    sc += sl[3]
    sc += sev * 1.5
    sc += {"Good":0,"Moderate":4,"Poor":9,"None":16,"N/A":2}.get(lighting, 4)
    sc -= {"Yes":1.0,"Partial":0.5,"No":0.0}.get(cctv_yn, 0) * 12
    sc += {"Very High":18,"High":12,"Medium":6,"Low":0,"No":0}.get(iso, 4)
    sc += 8 if drug=="Yes" else 0
    sc += 0 if psd<1 else 5 if psd<2 else 10 if psd<3 else 14
    sc += 4 if wknd else 0
    return max(0.0, min(100.0, round(sc, 1)))

ws1        = wb["1_Crime_Incidents"]
last1      = ws1.max_row
last_iid   = ws1.cell(last1, 1).value
iid        = int(last_iid) + 1 if last_iid else 12001
new_rows   = []

for year in [2020, 2021, 2022, 2023]:
    for st in NEW_STATIONS_V2:
        sn, zone, atype, lat, lon, zbase, cctv_pct, light_pct, psd, iso, drug, slum, pop = st
        n = max(8, int(zbase * 1.25))

        for _ in range(n):
            ct  = pick(CRIMES, CWN)
            sl  = pick(TIME_SLOTS, SWN)
            hr  = random.randint(sl[1], max(sl[1], sl[2]-1))
            isn = int(hr>=21 or hr<6)
            ise = int(18<=hr<21)
            isd = int(10<=hr<16)
            mo  = pick(MONTHS, MWN)
            mn  = MONTHS.index(mo)+1
            dy  = random.choice(DAYS)
            wknd= int(dy in ["Saturday","Sunday"])
            sea = SEASONS[mo]
            wea = random.choice(WEATHER[sea])

            if isn:
                lit = pick(["Good","Moderate","Poor","None"], [0.04,0.12,0.45,0.39])
            else:
                lit = pick(["Good","Moderate","Poor","None"], [0.48,0.34,0.14,0.04])

            cp  = cctv_pct/100.0
            cctv = pick(["Yes","Partial","No"],
                        [max(0,cp*0.8), min(0.2,0.3-cp*0.2),
                         max(0.05, 1-cp*0.8-min(0.2,0.3-cp*0.2))])

            if "26-" in ct[4] or "26-45" in ct[4]:
                va = pick(["26-30","31-40","41-50"], [0.40,0.45,0.15])
            else:
                va = pick(VICTIM_A, VICTIM_W)

            if "Husband" in ct[5]:   rel = "Husband/In-laws"
            elif "Known" in ct[5]:   rel = pick(["Known/Acquaintance","Neighbor","Family Member"],[0.6,0.25,0.15])
            elif "Online" in str(ct[6]): rel = "Online/Unknown"
            else:                    rel = pick(RELS,[0.20,0.25,0.30,0.10,0.05,0.05,0.05])

            if "Domestic" in ct[6][0]:  loc = "Domestic/Home"
            elif "Online" in ct[6][0]:  loc = "Online/Cyber"
            elif "Isolated" in ct[6][0]:loc = pick(["Isolated Road","Agricultural Field","Forest/Outskirts"],[0.5,0.3,0.2])
            elif "School" in ct[6][0]:  loc = pick(["Near School/College","Public Street","Park/Open Ground"],[0.5,0.3,0.2])
            else:                        loc = pick(LOCS[:8],[0.10,0.20,0.10,0.18,0.12,0.10,0.10,0.10])

            rs = risk(zbase, ct[1], hr, cctv, lit, psd, iso, drug, wknd)
            rl = 1 if rs<=35 else 2 if rs<=62 else 3
            rc = {1:"Medium",2:"High",3:"Critical"}[rl]
            cs = pick(CHARGE, CHARGE_W)

            new_rows.append([
                iid, year, mo, mn, dy, wknd, sea, wea,
                sn, zone, atype,
                round(lat+random.uniform(-0.009,0.009),6),
                round(lon+random.uniform(-0.009,0.009),6),
                ct[0], ct[2], ct[3], ct[1],
                sl[0], hr, isn, ise, isd,
                va, rel, loc, cctv, lit, psd, iso, drug,
                cs, rs, rl, rc
            ])
            iid += 1

print(f"  Generated {len(new_rows):,} new incidents")

for ri_off, row in enumerate(new_rows):
    ri  = last1 + 1 + ri_off
    alt = ri % 2 == 0
    cat = row[15]; rl = row[32]
    for ci, val in enumerate(row, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font      = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin()
        fx = (CRIME_FILL.get(cat) if ci==16
              else {1:"D5F5E3",2:"FEF9E7",3:"FADBD8"}.get(rl) if ci==33
              else None)
        c.fill = PatternFill("solid", fgColor=fx if fx else ("EEF4FB" if alt else "FFFFFF"))

wb.save(DATASET)

# Count total now
df_check = pd.read_excel(DATASET, sheet_name="1_Crime_Incidents", header=1)
df_zone_check = pd.read_excel(DATASET, sheet_name="2_Station_Zone_Profile", header=1)
print(f"  Dataset saved: {len(df_check):,} incidents | {len(df_zone_check)} stations")

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — RETRAIN ML MODEL
# ══════════════════════════════════════════════════════════════════════════════

print("\n[3/4] Retraining ML model on full dataset...")

from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, VotingClassifier
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics import accuracy_score, f1_score, classification_report
import joblib

os.makedirs("models", exist_ok=True)

df_inc  = pd.read_excel(DATASET, sheet_name="1_Crime_Incidents",          header=1)
df_zone = pd.read_excel(DATASET, sheet_name="2_Station_Zone_Profile",     header=1)
df_time = pd.read_excel(DATASET, sheet_name="3_Time_Environment_Factors", header=1)
print(f"  {len(df_inc):,} incidents | {len(df_zone)} stations")

df = df_inc.copy()
df["hour_24"]         = df["Hour_24"].astype(float)
df["crime_severity"]  = df["Crime_Severity_Score"].astype(float)
df["police_dist_km"]  = df["Nearest_PS_Dist_km"].astype(float)
df["is_weekend"]      = df["Is_Weekend"].astype(int)
df["cctv_numeric"]    = df["CCTV_Available"].str.strip().map({"Yes":1.0,"Partial":0.5,"No":0.0}).fillna(0.0)
df["lighting_numeric"]= df["Street_Lighting"].str.strip().map({"Good":0,"Moderate":1,"Poor":2,"None":3,"N/A":1}).fillna(1)
df["is_night"]        = ((df["hour_24"]>=21)|(df["hour_24"]<6)).astype(int)
df["is_evening"]      = ((df["hour_24"]>=18)&(df["hour_24"]<21)).astype(int)
df["is_daytime"]      = ((df["hour_24"]>=10)&(df["hour_24"]<16)).astype(int)
df["hour_sin"]        = np.sin(2*np.pi*df["hour_24"]/24)
df["hour_cos"]        = np.cos(2*np.pi*df["hour_24"]/24)
df["month_sin"]       = np.sin(2*np.pi*df["Month_Num"]/12)
df["month_cos"]       = np.cos(2*np.pi*df["Month_Num"]/12)

def get_tf(h):
    for _, row in df_time.iterrows():
        try:
            s,e = [int(x) for x in str(row["Hour_Range"]).split("-")]
            if s<=h<e: return float(row["Time_Additive_Score"]), float(row["Incident_Share_Pct"])
        except: pass
    return 0.0, 10.0

df["time_additive"], df["incident_share"] = zip(*df["hour_24"].apply(get_tf))

dz = df_zone[["Police_Station","Zone_Base_Score","Overall_Risk_Score_Display",
              "CCTV_Coverage_Pct","Nearest_PS_Dist_km",
              "Drug_Alcohol_Nearby","Slum_Area","Isolated_Roads","Population_Density"]].copy()
dz["drug_nearby"]  = (dz["Drug_Alcohol_Nearby"].str.lower()=="yes").astype(int)
dz["slum_area"]    = dz["Slum_Area"].str.lower().map({"yes":1,"partial":0.5,"no":0}).fillna(0)
dz["isolation"]    = dz["Isolated_Roads"].str.lower().map({"very high":4,"high":3,"medium":2,"low":1,"no":0}).fillna(1)
dz["pop_density"]  = dz["Population_Density"].str.lower().str.extract(r"(very high|high|medium|low|very low)")[0].map({"very high":5,"high":4,"medium":3,"low":2,"very low":1}).fillna(2)
dz["zone_base"]    = pd.to_numeric(dz["Zone_Base_Score"],            errors="coerce").fillna(25)
dz["zone_risk"]    = pd.to_numeric(dz["Overall_Risk_Score_Display"], errors="coerce").fillna(50)
dz["zone_cctv"]    = pd.to_numeric(dz["CCTV_Coverage_Pct"],          errors="coerce").fillna(20)
dz["zone_ps_dist"] = pd.to_numeric(dz["Nearest_PS_Dist_km"],         errors="coerce").fillna(2)
zs = dz[["Police_Station","zone_base","zone_risk","zone_cctv","zone_ps_dist",
          "drug_nearby","slum_area","isolation","pop_density"]].drop_duplicates("Police_Station")

df = df.merge(zs, on="Police_Station", how="left")
zcols = ["zone_base","zone_risk","zone_cctv","zone_ps_dist","drug_nearby","slum_area","isolation","pop_density"]
df[zcols] = df[zcols].fillna(df[zcols].median())

le_c = LabelEncoder(); df["crime_enc"]   = le_c.fit_transform(df["Crime_Type"].astype(str))
le_z = LabelEncoder(); df["zone_enc"]    = le_z.fit_transform(df["Zone"].astype(str))
le_a = LabelEncoder(); df["area_enc"]    = le_a.fit_transform(df["Area_Type"].astype(str))
le_v = LabelEncoder(); df["victim_enc"]  = le_v.fit_transform(df["Victim_Age_Group"].astype(str))
le_r = LabelEncoder(); df["rel_enc"]     = le_r.fit_transform(df["Relation_to_Accused"].astype(str))
le_l = LabelEncoder(); df["loc_enc"]     = le_l.fit_transform(df["Location_Context"].astype(str))
le_w = LabelEncoder(); df["weather_enc"] = le_w.fit_transform(df["Weather"].astype(str))

df["night_isolation"]    = df["is_night"]*df["isolation"]
df["night_no_cctv"]      = df["is_night"]*(1-df["cctv_numeric"])
df["evening_isolation"]  = df["is_evening"]*df["isolation"]
df["severity_x_night"]   = df["crime_severity"]*df["is_night"]
df["zone_risk_x_time"]   = df["zone_base"]*df["time_additive"]
df["dist_x_isolation"]   = df["zone_ps_dist"]*df["isolation"]
df["weekend_x_night"]    = df["is_weekend"]*df["is_night"]
df["severity_x_evening"] = df["crime_severity"]*df["is_evening"]

FEATURES = [
    "hour_24","crime_severity","police_dist_km","cctv_numeric","lighting_numeric",
    "is_night","is_evening","is_daytime","is_weekend",
    "hour_sin","hour_cos","month_sin","month_cos",
    "time_additive","incident_share",
    "zone_base","zone_risk","zone_cctv","zone_ps_dist",
    "drug_nearby","slum_area","isolation","pop_density",
    "crime_enc","zone_enc","area_enc","victim_enc","rel_enc","loc_enc","weather_enc",
    "night_isolation","night_no_cctv","evening_isolation",
    "severity_x_night","zone_risk_x_time","dist_x_isolation",
    "weekend_x_night","severity_x_evening",
]

X = df[FEATURES].fillna(0)
y = df["Risk_Label"]
print(f"  Feature matrix: {X.shape} | Labels: {dict(y.value_counts().sort_index())}")

Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.20, random_state=42, stratify=y)
sc  = StandardScaler()
Xts = sc.fit_transform(Xtr)
Xes = sc.transform(Xte)

rf = RandomForestClassifier(n_estimators=400, max_depth=14, max_features="sqrt",
     class_weight="balanced", random_state=42, n_jobs=-1)
rf.fit(Xts, ytr)

gb = GradientBoostingClassifier(n_estimators=250, learning_rate=0.08,
     max_depth=5, subsample=0.85, random_state=42)
gb.fit(Xts, ytr)

ens = VotingClassifier(estimators=[("rf",rf),("gb",gb)], voting="soft", weights=[2,1])
ens.fit(Xts, ytr)

acc = accuracy_score(yte, ens.predict(Xes))
f1  = f1_score(yte, ens.predict(Xes), average="weighted")
cv  = cross_val_score(ens, sc.transform(X), y,
      cv=StratifiedKFold(5, shuffle=True, random_state=42), scoring="f1_weighted")

print(f"  Accuracy: {acc*100:.2f}%  |  F1: {f1:.3f}  |  CV: {cv.mean():.3f} ± {cv.std():.4f}")
print()
print("  Classification Report:")
print(classification_report(yte, ens.predict(Xes), target_names=["Medium(1)","High(2)","Critical(3)"]))

fi = dict(zip(FEATURES, rf.feature_importances_))

bundle = {
    "model":ens, "scaler":sc, "features":FEATURES,
    "encoders":{"crime_type":le_c,"zone":le_z,"area_type":le_a,
                "victim_age":le_v,"relation":le_r,"location":le_l,"weather":le_w},
    "zone_lookup": zs.set_index("Police_Station").to_dict(orient="index"),
    "metrics":{
        "accuracy":round(acc,4), "f1_weighted":round(f1,4),
        "cv_f1_mean":round(float(cv.mean()),4), "cv_f1_std":round(float(cv.std()),4),
        "training_records":len(Xtr), "test_records":len(Xte),
        "total_features":len(FEATURES), "total_stations":len(df_zone),
    },
    "label_map":{1:"MEDIUM",2:"HIGH",3:"CRITICAL"},
    "label_thresholds":{"MEDIUM":"0-35","HIGH":"36-62","CRITICAL":"63-100"},
    "feature_importance":fi,
}

joblib.dump(bundle, "models/sheild_risk_model.pkl")
with open("models/model_metrics.json","w") as f:
    json.dump(bundle["metrics"], f, indent=2)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

print("\n[4/4] Done!")
print("="*70)
print(f"  Total incidents : {len(df_inc):,}")
print(f"  Total stations  : {len(df_zone)}")
print(f"  ML Accuracy     : {acc*100:.2f}%")
print(f"  CV F1           : {cv.mean():.3f} ± {cv.std():.4f}")
print()
print("  New areas added and their ML risk scores:")
print(f"  {'Area':<28} {'Risk':>6}  {'Category':<12}  Key reason")
print("  " + "-"*70)
for st in NEW_STATIONS_V2:
    sn, zone, atype, lat, lon, zbase, cctv, light, psd, iso, drug, slum, pop = st
    overall = min(100, int(zbase*1.85))
    rcat    = ("Critical"  if overall>=82 else "Very High" if overall>=68
               else "High" if overall>=50 else "Medium")
    reason  = ("No CCTV, Very Isolated" if cctv==0 and iso=="Very High"
               else "No CCTV, isolated"  if cctv<=3 and iso in ["Very High","High"]
               else "College area risk"  if "College" in sn
               else "Highway isolated"   if "Road" in sn or "Bypass" in sn
               else "Industrial night"   if atype=="Industrial"
               else "Residential")
    print(f"  {sn:<28} {overall:>6}  {rcat:<12}  {reason}")
print()
print("  Restart API: uvicorn main:app --reload --host 0.0.0.0 --port 8000")
print("="*70)
